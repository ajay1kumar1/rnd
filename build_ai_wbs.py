import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ──────────────────────────────────────────────────────────────
C_NAVY      = "0D1B2A"
C_DEEP_BLUE = "1B3A5C"
C_MID_BLUE  = "1F618D"
C_TEAL      = "148F77"
C_PURPLE    = "6C3483"
C_ORANGE    = "D35400"
C_WHITE     = "FFFFFF"
C_SILVER    = "F0F4F8"
C_PALE_BLUE = "D6EAF8"
C_PALE_GRN  = "D5F5E3"
C_PALE_PUR  = "E8DAEF"
C_GREY_TXT  = "2C3E50"
C_ACCENT    = "2980B9"

def fill(h): return PatternFill("solid", fgColor=h)
def bdr(style="thin", color="BFC9CA"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def fnt(bold=False, color="000000", size=9, italic=False, name="Calibri"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)
def aln(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════
#  SHEET 1 — WBS & ESTIMATION (AI only)
# ═══════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "WBS & Estimation"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A8"

# column widths
CW = {"A":7,"B":10,"C":30,"D":30,"E":16,"F":10,"G":9,
      "H":13,"I":13,"J":13,"K":15,"L":11,"M":11,"N":10,"O":20}
for col,w in CW.items(): ws.column_dimensions[col].width = w

# ── Banner ──
for r,h in [(1,8),(2,38),(3,18),(4,8),(5,14),(6,8)]: ws.row_dimensions[r].height=h

ws.merge_cells("A1:O1"); ws["A1"].fill = fill(C_NAVY)

ws.merge_cells("A2:O2")
ws["A2"] = "EXPERIAN · AI ENGINEERING PHASE 1  —  WORK BREAKDOWN STRUCTURE & EFFORT ESTIMATION"
ws["A2"].font = Font(name="Calibri",bold=True,color=C_WHITE,size=15)
ws["A2"].fill = fill(C_NAVY); ws["A2"].alignment = aln("center","center")

ws.merge_cells("A3:O3")
ws["A3"] = ("Phase 1  |  Intelligent Search · Conversational Chatbot · DocuSafe AI Insights · "
            "Report Recommendations · Knowledge Mgmt / Help · AEO / AI Readability")
ws["A3"].font = Font(name="Calibri",italic=True,color="A9CCE3",size=9)
ws["A3"].fill = fill(C_NAVY); ws["A3"].alignment = aln("center","center")

ws.merge_cells("A4:O4"); ws["A4"].fill = fill(C_TEAL)
for col in range(2,16): ws.cell(2,col).fill=fill(C_NAVY); ws.cell(3,col).fill=fill(C_NAVY); ws.cell(4,col).fill=fill(C_TEAL)

# meta strip
ws.row_dimensions[5].height = 15
ws.merge_cells("A5:G5"); ws["A5"]="  Client: Experian  |  Prepared by: AI Engineering Team  |  Date: June 2026"; ws["A5"].font=fnt(italic=True,color=C_GREY_TXT,size=8); ws["A5"].fill=fill(C_SILVER); ws["A5"].alignment=aln(indent=1)
ws.merge_cells("H5:K5"); ws["H5"]="1 Person-Day = 8 hrs  |  1 PM = 21.5 days = 172 hrs"; ws["H5"].font=fnt(italic=True,color=C_GREY_TXT,size=8); ws["H5"].fill=fill(C_SILVER); ws["H5"].alignment=aln("center")
ws.merge_cells("L5:O5"); ws["L5"]="Gen AI uplift applied — ML/AI team only"; ws["L5"].font=fnt(italic=True,color=C_GREY_TXT,size=8); ws["L5"].fill=fill(C_SILVER); ws["L5"].alignment=aln("center")

ws.merge_cells("A6:O6"); ws["A6"].fill=fill(C_NAVY)

# ── Column headers ──
ws.row_dimensions[7].height = 36
HEADERS = ["WBS ID","Phase","Task / AI Component","Description / Scope",
           "AI Layer","Component Type","Size",
           "ML/AI Eng\n(Person Days)","Data Eng\n(Person Days)",
           "MLOps\n(Person Days)","Total Effort\n(Person Days)",
           "Gen AI\nMin (PD)","Gen AI\nMax (PD)","Priority","Notes / Assumptions"]
for ci,h in enumerate(HEADERS,1):
    c=ws.cell(7,ci,h)
    c.font=Font(name="Calibri",bold=True,color=C_WHITE,size=9)
    c.fill=fill(C_DEEP_BLUE); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=bdr()

# ── WBS DATA ─────────────────────────────────────────────────────────
# (wbs, phase, task, description, ai_layer, comp_type, size, ml_ai, data_eng, mlops, priority, notes)
TASKS = [
    # ── 1. INTELLIGENT SEARCH ────────────────────────────────────────
    {"S":"1.0  INTELLIGENT SEARCH — Entity Search Assistance & Report Routing"},
    {"s":"1.1  Intent Classification & Orchestration"},
    ("1.1.1","P1","Intent Classifier — Search","Fine-tune/prompt-engineer intent classifier to detect SEARCH intent from NL queries","Orchestration","LLM/Agent","M",8,2,2,"High","LangGraph; Bedrock Anthropic as primary"),
    ("1.1.2","P1","Orchestrator Routing Logic","LangGraph stateful graph: route classified intents to Search agent, handle fallback","Orchestration","LLM/Agent","M",6,1,2,"High","Fargate ECS; Python/Django"),
    ("1.1.3","P1","Prompt Store — Search Templates","Author, version, and deploy prompt templates for entity search use-case","Orchestration","Prompt Eng","S",5,0,1,"High","MCP Prompt Store; A/B versioning"),
    {"s":"1.2  Retrieval & Vector Store"},
    ("1.2.1","P1","Document Chunking Pipeline","Strategy: sliding window + semantic chunking; metadata tagging (BU, doc-type, recency)","RAG","Data Pipeline","L",6,8,2,"High","AWS Glue / Lambda; S3 input"),
    ("1.2.2","P1","Embedding Model Setup","Configure & benchmark CLIP / E5 / BGE embeddings on SageMaker endpoint","RAG","Embeddings","L",8,4,3,"High","SageMaker Inf2; ECR model registry"),
    ("1.2.3","P1","OpenSearch Index Design","Schema: dense vector (k-NN) + sparse BM25 hybrid; ACL metadata fields","RAG","Vector Store","M",5,3,2,"High","OpenSearch Serverless or managed"),
    ("1.2.4","P1","Hybrid Search Implementation","Dense + sparse retrieval fusion; score normalisation before reranking","RAG","Vector Store","M",6,2,2,"High","Reciprocal Rank Fusion (RRF)"),
    ("1.2.5","P1","Auto Re-ingest Hook","If doc exists but not indexed → enqueue S3 event → Lambda → re-embed pipeline","RAG","Data Pipeline","M",5,4,2,"Medium","Dead-letter queue on failure"),
    {"s":"1.3  Reranking & Confidence Gate"},
    ("1.3.1","P1","Cross-Encoder Reranker","Deploy cross-encoder rerank model; score top-k candidates from retriever","RAG","Reranker","M",8,2,2,"High","SageMaker endpoint; ms-marco-MiniLM"),
    ("1.3.2","P1","Confidence Threshold Gate","Configure threshold; route low-confidence → fallback LLM long-context summarisation","RAG","Guardrails","S",5,1,1,"High","Threshold tunable per use-case"),
    {"s":"1.4  MCP Tools & Citation Builder"},
    ("1.4.1","P1","MCP Server — Entity Search Tool","search_entities MCP tool: ACL-aware query to Aurora entity records","MCP/Tools","MCP Server","M",8,3,2,"High","Aurora + pgvector; PrivateLink"),
    ("1.4.2","P1","MCP Server — ACL Check Tool","Validate user permissions against document ACL before returning results","MCP/Tools","MCP Server","S",5,1,1,"High","IAM + DynamoDB ACL store"),
    ("1.4.3","P1","Citation Builder Service","Attach source doc metadata, chunk references, and confidence scores to response","MCP/Tools","MCP Server","S",5,1,1,"High","Response grounding per Bedrock standard"),
    {"s":"1.5  PII & Security"},
    ("1.5.1","P1","PII Tokenisation Vault — MY","Lambda tokeniser: detect (Comprehend NER) → replace with PERSON_/EMAIL_/IC_ tokens","PII/Security","PII Vault","L",8,3,3,"High","DynamoDB vault; KMS CMK; stays MY-resident"),
    ("1.5.2","P1","De-tokenisation on Return Path","Route model response back through MY vault; swap tokens → PII before user egress","PII/Security","PII Vault","M",6,2,2,"High","Fail-closed by default (Q26 compliance)"),
    ("1.5.3","P1","Bedrock Native Guardrails Config","Topic policy, PII filter, toxicity filter per Intelligent Search use-case","PII/Security","Guardrails","S",4,0,1,"High","Bedrock Guardrails; DynamoGuard optional"),
    # ── 2. CONVERSATIONAL CHATBOT ─────────────────────────────────────
    {"S":"2.0  CONVERSATIONAL CHATBOT — In-Portal Workflow Guidance"},
    {"s":"2.1  Conversation State & Context Management"},
    ("2.1.1","P1","Conversation State Store","Design DynamoDB conversation state schema: session_id, message history, user context, BU","Orchestration","LLM/Agent","M",6,3,2,"High","TTL-based session expiry"),
    ("2.1.2","P1","Intent Classifier — Chat","Fine-tune classifier for chat intents: workflow_help, error_resolution, escalation","Orchestration","LLM/Agent","M",8,2,2,"High","Shared orchestrator; separate prompt policy"),
    ("2.1.3","P1","Multi-turn Dialogue Manager","Manage conversation turns; inject context window; handle user corrections","Orchestration","LLM/Agent","L",10,2,3,"High","LangGraph stateful graph; sliding context"),
    {"s":"2.2  RAG for Workflow Knowledge"},
    ("2.2.1","P1","Workflow Knowledge Base Ingestion","Ingest BU workflow docs, SOPs, and error codes; chunk & embed into OpenSearch","RAG","Data Pipeline","L",5,8,2,"High","Shared vector store — namespace partitioned"),
    ("2.2.2","P1","Contextual Retrieval (Chat)","Retrieve relevant workflow steps based on current conversation step + user context","RAG","Vector Store","M",6,2,2,"High","Hybrid search + rerank shared with Search"),
    ("2.2.3","P1","Step-by-Step Response Composer","Format grounded responses as numbered steps; attach citation links","RAG","LLM/Agent","M",6,1,2,"High","Claude / GPT-4V long-context LLM"),
    {"s":"2.3  Guardrails & Safety"},
    ("2.3.1","P1","Chat-Specific Guardrails","PII / toxicity / topic policy scoped to chatbot use-case in Bedrock Guardrails","PII/Security","Guardrails","S",4,0,1,"High","Separate policy from Search"),
    ("2.3.2","P1","Escalation Trigger Detection","Detect low-confidence / out-of-scope → escalation flag → human handoff signal","Orchestration","LLM/Agent","S",5,1,1,"Medium","Webhook to BU portal CRM"),
    # ── 3. DOCUSAFE AI INSIGHTS ──────────────────────────────────────
    {"S":"3.0  DOCUSAFE AI INSIGHTS — Document & Portfolio Risk Insights"},
    {"s":"3.1  Multimodal Document Ingestion"},
    ("3.1.1","P1","Multimodal Document Parser","Parse PDF, image, table inputs; extract text + layout; detect doc type","RAG","Data Pipeline","L",8,6,3,"High","AWS Textract + custom post-processor"),
    ("3.1.2","P1","Multimodal Encoder Pipeline","CLIP (images) + E5/BGE (text) dual-encoder; store embeddings with doc metadata","RAG","Embeddings","L",10,5,3,"High","SageMaker multi-container endpoint"),
    ("3.1.3","P1","Portfolio-level Ingestion Orchestration","Orchestrate batch ingestion for analyst portfolio: progress tracking, partial re-ingest","RAG","Data Pipeline","M",6,4,2,"High","Step Functions + SQS; S3 event triggers"),
    {"s":"3.2  Risk Insight Generation"},
    ("3.2.1","P1","Planner Agent — DocuSafe","Decompose analyst query into sub-queries across doc set; multi-step plan","Orchestration","LLM/Agent","L",10,2,3,"High","LangGraph planner node; tool-calling loop"),
    ("3.2.2","P1","Query Rewriter & Multi-Query Expansion","Expand single query to multi-perspective sub-queries to improve recall","RAG","LLM/Agent","M",6,1,2,"High","HyDE / multi-query strategy"),
    ("3.2.3","P1","Knowledge Graph — Entity Links","Build entity relationship graph across portfolio docs; link related entities","RAG","Data Pipeline","XL",12,10,4,"High","Neptune or OpenSearch graph index"),
    ("3.2.4","P1","Risk Insight Composer (Grounded)","Generate risk narratives with inline citations from retrieved doc chunks","RAG","LLM/Agent","L",8,2,2,"High","Claude Opus / GPT-4V multimodal LLM"),
    {"s":"3.3  Validation & Self-Reflection"},
    ("3.3.1","P1","Self-Reflection Validation Loop","Post-generation: verify groundedness; hallucination check; quality score","Orchestration","LLM/Agent","M",8,1,2,"High","Reflective validation node in LangGraph"),
    ("3.3.2","P1","Quality Scorer — Hallucination Detection","Score response on faithfulness, relevance, context precision (RAGAS metrics)","Orchestration","LLM/Agent","M",8,2,2,"High","RAGAS + custom scorer Lambda"),
    ("3.3.3","P1","Decision Gate — Accept / Retry / Refine","Route low-quality responses to retry/refine loop; cap at 3 iterations","Orchestration","LLM/Agent","S",5,0,1,"High","Max-retry guard to prevent infinite loops"),
    # ── 4. REPORT RECOMMENDATIONS ─────────────────────────────────────
    {"S":"4.0  REPORT RECOMMENDATIONS — Context-Aware Credit Report Selection"},
    {"s":"4.1  Report Classification & Routing"},
    ("4.1.1","P1","MCP Server — Classify Report","classify_report MCP tool: Bedrock classifier maps user context → product code","MCP/Tools","MCP Server","M",8,2,2,"High","Bedrock classifier model; Aurora product table"),
    ("4.1.2","P1","MCP Server — Route Report","route_report MCP tool: select ranked report options with business rule overlay","MCP/Tools","MCP Server","M",8,2,2,"High","Step Functions routing logic"),
    ("4.1.3","P1","MCP Server — Audit Log","log_decision MCP tool: structured audit record → DynamoDB + S3 archive","MCP/Tools","MCP Server","S",5,2,1,"High","Immutable S3 + DynamoDB TTL for hot queries"),
    {"s":"4.2  Contextual Retrieval & Generation"},
    ("4.2.1","P1","Vector Store — Final Results Index","Separate OpenSearch index for ranked credit report selections + citation metadata","RAG","Vector Store","M",5,3,2,"High","Namespace-partitioned shared cluster"),
    ("4.2.2","P1","Report Generation LLM","Generate personalised report recommendation narrative; attach ranked options","RAG","LLM/Agent","M",6,1,2,"High","Claude / GPT-4 report-gen prompt"),
    ("4.2.3","P1","Confidence Threshold & Rerank","Threshold gate scoped to Report Recommendations; rerank on business rules + semantic score","RAG","Reranker","S",5,1,1,"High","Shared cross-encoder; custom rule overlay"),
    # ── 5. KNOWLEDGE MANAGEMENT / HELP ───────────────────────────────
    {"S":"5.0  KNOWLEDGE MANAGEMENT & HELP — Contextual FAQ & Error Prevention"},
    {"s":"5.1  KM Content Ingestion"},
    ("5.1.1","P1","KM Document Ingestion Pipeline","Ingest FAQ docs, error-resolution guides, policy docs; classify internal vs external","RAG","Data Pipeline","M",5,6,2,"High","Internal-only flag in OpenSearch ACL metadata"),
    ("5.1.2","P1","Context & Document Status Tagger","Tag doc visibility: internal-only / public; inject into ACL field at chunk level","RAG","Data Pipeline","S",4,3,1,"High","Lambda metadata enrichment step"),
    {"s":"5.2  FAQ & Error Resolution Generation"},
    ("5.2.1","P1","Contextual FAQ Retriever","Retrieve relevant FAQ chunks based on user error code / support query","RAG","Vector Store","M",6,2,2,"High","Shared OpenSearch cluster; KM namespace"),
    ("5.2.2","P1","Direct Error-Resolution Composer","Generate step-by-step error resolution summary with preventative action steps","RAG","LLM/Agent","M",6,1,2,"High","Short-context LLM; sub-500ms SLA"),
    ("5.2.3","P1","Guardrails — KM Use-Case","Scoped PII / topic policy for KM; block internal-only content to external users","PII/Security","Guardrails","S",4,0,1,"High","Content visibility enforced at retrieval + guardrail"),
    # ── 6. AEO / AI READABILITY ───────────────────────────────────────
    {"S":"6.0  AEO / AI READABILITY — Generative Engine Optimisation"},
    {"s":"6.1  Content Pre-Processing"},
    ("6.1.1","P1","SEO & Structure Pre-Processor","Parse human-generated content; enforce heading hierarchy, schema.org tags, structured data","AEO","Data Pipeline","M",4,4,2,"Medium","Lambda; NLP-based structural analysis"),
    ("6.1.2","P1","PII Tokenisation — AEO Content","Detect and tokenise PII in content before LLM optimisation pass","PII/Security","PII Vault","S",4,2,1,"High","Shared MY vault; PERSON_/EMAIL_ tokens"),
    {"s":"6.2  AI Readability Optimisation"},
    ("6.2.1","P1","LLM Content Optimiser","Rewrite content for AI readability: answer-first structure, concise factual statements","AEO","LLM/Agent","M",6,1,2,"Medium","Claude Sonnet; custom AEO prompt template"),
    ("6.2.2","P1","Vector Store — Schema-Tagged Content","Index optimised content with schema tags; enable semantic search by AEO consumers","AEO","Vector Store","M",4,3,2,"Medium","OpenSearch; AEO namespace"),
    ("6.2.3","P1","Guardrails — AEO Content Policy","Enforce content policy: no PII leakage, toxicity filter, factual consistency check","PII/Security","Guardrails","S",3,0,1,"Medium","Bedrock Guardrails; AEO topic policy"),
    # ── 7. AI GATEWAY & SHARED INFRASTRUCTURE ─────────────────────────
    {"S":"7.0  AI GATEWAY, SHARED INFRASTRUCTURE & CONTINUOUS LEARNING"},
    {"s":"7.1  AI Gateway (LiteLLM / Ascend One)"},
    ("7.1.1","P1","AI Gateway Setup — Per Use-Case Rate Limits","Configure per-use-case rate limits, token budgets, and timeout policies","AI Gateway","Platform","M",6,2,2,"High","LiteLLM / Ascend One AI Gateway"),
    ("7.1.2","P1","Ordered Fallback Chain","Primary Bedrock → SageMaker OSS → Azure OpenAI fallback; health-check routing","AI Gateway","Platform","M",8,1,2,"High","Bedrock Anthropic → Databricks OSS → Gemini"),
    ("7.1.3","P1","Semantic Cache Layer","Cache frequent query→response pairs; cosine similarity threshold for cache hit","AI Gateway","Platform","M",6,2,2,"High","Redis / OpenSearch semantic cache"),
    ("7.1.4","P1","DynamoGuard Integration","In/out policy enforcement via DynamoGuard where required by use-case","AI Gateway","Platform","S",4,1,1,"Medium","Optional per use-case; Bedrock Guardrails default"),
    {"s":"7.2  Audit, Metering & Observability"},
    ("7.2.1","P1","Kafka Audit Stream Setup","AI Gateway → Kafka topic per use-case; structured audit event schema","Audit/Obs","Platform","M",4,6,2,"High","MSK Kafka; S3 raw audit sink"),
    ("7.2.2","P1","ETL — Audit to Inference/Metering Store","Glue ETL: S3 raw audit → S3 inference + metering store (FinOps + IR)","Audit/Obs","Data Pipeline","M",2,8,2,"High","Glue + Athena; cost attribution per use-case"),
    ("7.2.3","P1","Splunk Log Integration","Route PII-free logs to Splunk (SG); MY-resident Splunk index for PII-bearing lines","Audit/Obs","Platform","S",3,2,1,"High","Sanitised audit only crosses MY→SG border"),
    ("7.2.4","P1","CloudWatch Dashboards & Alarms","LLM latency p95/p99, cache hit rate, hallucination score, token spend per use-case","Audit/Obs","Platform","M",3,2,2,"High","CloudWatch + PagerDuty alerting"),
    {"s":"7.3  Continuous Learning Loop"},
    ("7.3.1","P1","Telemetry Capture Service","Capture user signals: thumbs-up/down, dwell time, correction signals; store in S3","ContLearning","ML Pipeline","M",6,4,2,"High","Lambda event collector; Kinesis firehose"),
    ("7.3.2","P1","Golden Dataset Curation Pipeline","Human-in-loop review of flagged responses; curate golden Q&A pairs for fine-tuning","ContLearning","ML Pipeline","L",10,5,3,"High","Label Studio / custom review UI on S3"),
    ("7.3.3","P1","SageMaker Fine-Tuning Pipeline","LoRA / QLoRA fine-tune open-source LLM on golden set; eval on held-out test","ContLearning","ML Pipeline","XL",12,6,4,"High","SageMaker Pipelines; ECR model registry"),
    ("7.3.4","P1","Model Evaluation & Promotion Gate","RAGAS + custom benchmark; A/B shadow deployment; promote if p99 latency & quality pass","ContLearning","ML Pipeline","L",10,3,3,"High","SageMaker Model Monitor; blue/green deploy"),
    {"s":"7.4  Re-embed & Re-index Pipeline"},
    ("7.4.1","P1","Re-embed Trigger & Scheduler","Scheduled + event-triggered re-embedding for updated docs; delta-only re-index","ContLearning","Data Pipeline","M",4,6,2,"High","EventBridge scheduler; SQS queue"),
    ("7.4.2","P1","Vector Store Refresh & Index Swap","Zero-downtime index alias swap after re-index; rollback capability","ContLearning","Data Pipeline","M",4,6,2,"High","OpenSearch index aliases"),
    # ── 8. PLATFORM INFRASTRUCTURE ────────────────────────────────────
    {"S":"8.0  PLATFORM INFRASTRUCTURE — AWS Cloud-Native"},
    {"s":"8.1  Network & Edge"},
    ("8.1.1","P1","VPC & Private Subnet Design","Customer-managed VPC; private subnets; VPC endpoints (PrivateLink) for all AWS services","Infra","Platform","L",2,2,6,"High","Data stays in-region; no public internet for LLM calls"),
    ("8.1.2","P1","Incapsula WAF / AWS WAF Config","OWASP Top-10 + bot filtering; TLS termination at edge; rate limiting rules","Infra","Platform","M",1,1,4,"High","Incapsula primary; AWS WAF on ALB"),
    ("8.1.3","P1","CloudFront + ALB Setup","CDN edge cache; ALB TLS termination; ECS Fargate target group config","Infra","Platform","M",1,1,4,"High","CloudFront distribution per environment"),
    {"s":"8.2  Compute & Serving"},
    ("8.2.1","P1","ECS Fargate — Orchestrator Service","Orchestrator + agent runtime containers; task definitions, auto-scaling, health checks","Infra","Platform","L",2,2,6,"High","Django/Python orchestrator; Fargate Spot for dev"),
    ("8.2.2","P1","SageMaker Endpoints — LLM Serving","Llama 3 / Mistral endpoints; Inf2 inference accelerators; multi-model endpoint","Infra","Platform","L",4,2,6,"High","ECR model registry; A/B endpoint variants"),
    ("8.2.3","P1","Lambda Functions — Glue & Tools","Tool functions: tokeniser, re-ingest, MCP glue; Lambda layers for shared deps","Infra","Platform","M",2,2,4,"High","Lambda SnapStart; VPC Lambda config"),
    {"s":"8.3  Security & Compliance"},
    ("8.3.1","P1","IAM Roles & Least-Privilege Policies","Service-to-service IAM roles; resource-based policies; SCPs for dev/staging/prod","Infra","Security","M",1,1,4,"High","IAM Access Analyser; no wildcard policies"),
    ("8.3.2","P1","KMS Customer-Managed Keys","CMK per data classification: vault, vector store, audit S3; key rotation policy","Infra","Security","M",1,1,3,"High","KMS + Secrets Manager integration"),
    ("8.3.3","P1","CloudTrail & Security Hub","API audit logging; Security Hub findings aggregation; GuardDuty threat detection","Infra","Security","S",1,1,2,"High","CIS AWS Foundations Benchmark"),
    # ── 9. NFR ────────────────────────────────────────────────────────
    {"S":"9.0  NON-FUNCTIONAL REQUIREMENTS & PERFORMANCE ENGINEERING"},
    {"s":"9.1  Performance & Latency SLAs"},
    ("9.1.1","P1","End-to-End Latency Optimisation","PII vault p99 < 50ms; total AI response p95 < 3s; semantic cache hit > 40%","NFR","Performance","L",8,4,4,"High","Latency budget: vault 50ms + gateway 100ms + LLM"),
    ("9.1.2","P1","Load & Stress Testing (AI Endpoints)","k6 load tests: SageMaker endpoints, AI Gateway, OpenSearch; scale-to-zero test","NFR","Performance","M",4,2,4,"High","Target: 200 rps sustained; 500 rps peak 1 min"),
    {"s":"9.2  Model Quality Metrics"},
    ("9.2.1","P1","RAGAS Evaluation Baseline","Establish faithfulness, answer relevance, context precision baselines for each use-case","NFR","Quality","M",8,2,2,"High","Minimum: faithfulness > 0.85; context precision > 0.80"),
    ("9.2.2","P1","Hallucination Rate Monitoring","Automated hallucination scoring on sampled production responses; alert on regression","NFR","Quality","M",6,2,2,"High","CloudWatch custom metric; PagerDuty on threshold breach"),
]

# ── Render rows ──────────────────────────────────────────────────────
row = 8
data_rows = []
shade = False

def section_row(r, label, bg):
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f"A{r}:O{r}")
    c = ws.cell(r, 1, label)
    c.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)
    c.fill = fill(bg); c.alignment = aln("left","center",indent=1); c.border = bdr()
    for col in range(2,16):
        ws.cell(r,col).fill=fill(bg); ws.cell(r,col).border=bdr()

def sub_row(r, label):
    ws.row_dimensions[r].height = 17
    ws.merge_cells(f"A{r}:O{r}")
    c = ws.cell(r, 1, label)
    c.font = Font(name="Calibri", bold=True, color=C_WHITE, size=9)
    c.fill = fill(C_TEAL); c.alignment = aln("left","center",indent=2); c.border = bdr()
    for col in range(2,16):
        ws.cell(r,col).fill=fill(C_TEAL); ws.cell(r,col).border=bdr()

for item in TASKS:
    if isinstance(item, dict):
        if "S" in item: section_row(row, item["S"], C_NAVY); row+=1
        elif "s" in item: sub_row(row, item["s"]); row+=1
        continue

    wbs_id,phase,task,desc,ai_layer,comp,size,ml,de,mlops,priority,notes = item
    ws.row_dimensions[row].height = 28
    bg = C_PALE_BLUE if shade else "FFFFFF"
    shade = not shade
    qa = round((ml+de+mlops)*0.20)
    pri_color = {"High":C_NAVY,"Medium":C_MID_BLUE,"Low":"808080"}.get(priority,C_GREY_TXT)

    vals = [wbs_id, phase, task, desc, ai_layer, comp, size, ml, de, mlops,
            f"=H{row}+I{row}+J{row}",
            f"=ROUND((H{row}+I{row})*0.15,1)",
            f"=ROUND((H{row}+I{row})*0.35,1)",
            priority, notes]
    aligns = [aln("center"), aln("center"),
              aln("left",wrap=True,indent=1), aln("left",wrap=True,indent=1),
              aln("center"), aln("center"), aln("center"),
              aln("center"), aln("center"), aln("center"),
              aln("center"), aln("center"), aln("center"),
              aln("center"), aln("left",wrap=True,indent=1)]

    for ci,val in enumerate(vals,1):
        c = ws.cell(row,ci,val)
        c.fill=fill(bg); c.border=bdr()
        c.alignment=aligns[ci-1]
        c.font=fnt(size=9,color=C_GREY_TXT)
        if ci in (8,9,10): c.font=fnt(size=9,color="154360")
        if ci==11: c.font=Font(name="Calibri",bold=True,size=9,color=C_NAVY)
        if ci==14: c.font=Font(name="Calibri",bold=True,size=9,color=pri_color)
    data_rows.append(row); row+=1

# Grand Total row
ws.row_dimensions[row].height = 22
ws.merge_cells(f"A{row}:G{row}")
gt = ws.cell(row,1,"  GRAND TOTAL EFFORT (Person Days)")
gt.font=Font(name="Calibri",bold=True,color=C_WHITE,size=10)
gt.fill=fill(C_NAVY); gt.alignment=aln("left","center",indent=1); gt.border=bdr()
for c in range(2,8): ws.cell(row,c).fill=fill(C_NAVY); ws.cell(row,c).border=bdr()

for ci,formula in enumerate([
    f"=SUM("+ ",".join(f"H{r}" for r in data_rows)+")",
    f"=SUM("+ ",".join(f"I{r}" for r in data_rows)+")",
    f"=SUM("+ ",".join(f"J{r}" for r in data_rows)+")",
    f"=SUM("+ ",".join(f"K{r}" for r in data_rows)+")",
    f"=SUM("+ ",".join(f"L{r}" for r in data_rows)+")",
    f"=SUM("+ ",".join(f"M{r}" for r in data_rows)+")",
    "",""],start=8):
    c=ws.cell(row,ci,formula)
    c.font=Font(name="Calibri",bold=True,color=C_WHITE,size=10)
    c.fill=fill(C_NAVY); c.alignment=aln("center"); c.border=bdr()

gt_row=row; row+=1
# PM row
ws.row_dimensions[row].height=18
ws.merge_cells(f"A{row}:G{row}")
pm=ws.cell(row,1,"  Total Effort (Person Months)  [1 PM = 21.5 days]")
pm.font=fnt(bold=True,color=C_DEEP_BLUE,size=9); pm.fill=fill(C_PALE_BLUE); pm.alignment=aln("left","center",indent=1); pm.border=bdr()
for c in range(2,8): ws.cell(row,c).fill=fill(C_PALE_BLUE); ws.cell(row,c).border=bdr()
for ci,col in enumerate(["H","I","J","K","L","M"],start=8):
    c=ws.cell(row,ci,f"=ROUND({col}{gt_row}/21.5,1)")
    c.font=fnt(bold=True,color=C_DEEP_BLUE,size=9); c.fill=fill(C_PALE_BLUE); c.alignment=aln("center"); c.border=bdr()
ws.cell(row,14).fill=fill(C_PALE_BLUE); ws.cell(row,14).border=bdr()
ws.cell(row,15).fill=fill(C_PALE_BLUE); ws.cell(row,15).border=bdr()

# ═══════════════════════════════════════════════════════════════════════
#  SHEET 2 — RESOURCE LOADING
# ═══════════════════════════════════════════════════════════════════════
rl = wb.create_sheet("Resource Loading (RL)")
rl.sheet_view.showGridLines = False
rl.freeze_panes = "C9"

rl.column_dimensions["A"].width = 30
rl.column_dimensions["B"].width = 12
for i in range(3,35): rl.column_dimensions[get_column_letter(i)].width = 5.5

# Title
for r,h in [(1,8),(2,36),(3,16),(4,8)]: rl.row_dimensions[r].height=h
rl.merge_cells("A2:AH2")
t=rl.cell(2,1,"EXPERIAN · AI ENGINEERING PHASE 1 — RESOURCE LOADING PLAN")
t.font=Font(name="Calibri",bold=True,color=C_WHITE,size=15); t.fill=fill(C_NAVY); t.alignment=aln("center","center")
rl.merge_cells("A3:AH3")
s=rl.cell(3,1,"AI-Only Team  |  6 Milestones  ·  24 Weeks  ·  Offshore Delivery  |  No FE / BE Developers")
s.font=Font(name="Calibri",italic=True,color="A9CCE3",size=10); s.fill=fill(C_NAVY); s.alignment=aln("center","center")
rl.merge_cells("A4:AH4"); rl.cell(4,1).fill=fill(C_TEAL)
for c in range(2,35): rl.cell(2,c).fill=fill(C_NAVY); rl.cell(3,c).fill=fill(C_NAVY); rl.cell(4,c).fill=fill(C_TEAL)

# Milestone bands (row 5)
MILESTONES = [
    ("M1","Platform Setup & RAG Foundations",2,C_NAVY),
    ("M2","Use-Case AI Dev Sprint 1\n(Search + Chatbot)",5,C_DEEP_BLUE),
    ("M3","Use-Case AI Dev Sprint 2\n(DocuSafe + Reports + KM)",5,C_MID_BLUE),
    ("M4","AEO + AI Gateway + Guardrails",4,"145A32"),
    ("M5","Cont. Learning + Infra Hardening",4,C_PURPLE),
    ("M6","Performance, Eval & Go-Live",4,C_ORANGE),
]
rl.row_dimensions[5].height=24; rl.row_dimensions[6].height=24; rl.row_dimensions[7].height=16; rl.row_dimensions[8].height=30

wk_col=3
for (m_name,m_desc,weeks,m_col) in MILESTONES:
    s_col=wk_col; e_col=wk_col+weeks-1
    rl.merge_cells(f"{get_column_letter(s_col)}5:{get_column_letter(e_col)}5")
    mc=rl.cell(5,s_col,f"{m_name}: {m_desc}")
    mc.font=Font(name="Calibri",bold=True,color=C_WHITE,size=8)
    mc.fill=fill(m_col); mc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); mc.border=bdr()
    for c in range(s_col+1,e_col+1):
        rl.cell(5,c).fill=fill(m_col); rl.cell(5,c).border=bdr()
    wk_col+=weeks

# Row 6 — Phase labels
wk_col=3
PHASE_LABELS=[("M1: Wks 1-2",2),("M2: Wks 3-7",5),("M3: Wks 8-12",5),("M4: Wks 13-16",4),("M5: Wks 17-20",4),("M6: Wks 21-24",4)]
for (pl,pw) in PHASE_LABELS:
    s_col=wk_col; e_col=wk_col+pw-1
    rl.merge_cells(f"{get_column_letter(s_col)}6:{get_column_letter(e_col)}6")
    c=rl.cell(6,s_col,pl)
    c.font=Font(name="Calibri",bold=True,color=C_WHITE,size=8); c.fill=fill(C_DEEP_BLUE)
    c.alignment=aln("center"); c.border=bdr()
    for cc in range(s_col+1,e_col+1):
        rl.cell(6,cc).fill=fill(C_DEEP_BLUE); rl.cell(6,cc).border=bdr()
    wk_col+=pw

# Row 7 — Effort unit label
rl.cell(7,1,"Resource / Role").font=Font(name="Calibri",bold=True,color=C_WHITE,size=9)
rl.cell(7,1).fill=fill(C_DEEP_BLUE); rl.cell(7,1).alignment=aln("center"); rl.cell(7,1).border=bdr()
rl.cell(7,2,"Location").font=Font(name="Calibri",bold=True,color=C_WHITE,size=9)
rl.cell(7,2).fill=fill(C_DEEP_BLUE); rl.cell(7,2).alignment=aln("center"); rl.cell(7,2).border=bdr()
for w in range(1,25):
    c=rl.cell(7,w+2,f"W{w}")
    c.font=Font(name="Calibri",bold=True,color=C_WHITE,size=8)
    c.fill=fill(C_DEEP_BLUE); c.alignment=aln("center"); c.border=bdr()

# Row 8 — effort legend
rl.merge_cells("A8:B8")
rl.cell(8,1,"HEADCOUNT per week  (number of people allocated to this programme)").font=fnt(italic=True,color=C_GREY_TXT,size=8)
rl.cell(8,1).fill=fill(C_SILVER); rl.cell(8,1).alignment=aln(indent=1); rl.cell(8,1).border=bdr()
for w in range(1,25):
    c=rl.cell(8,w+2)
    c.fill=fill(C_SILVER); c.border=bdr()

# ── RL Resource Data ─────────────────────────────────────────────────
# (role, location, [w1..w24])
POD_DATA = [
    ("AI LEADERSHIP & ARCHITECTURE", C_NAVY, [
        ("AI Engineering Manager",   "Offshore", [1]*24),
        ("AI Solution Architect",     "Offshore", [1]*24),
    ]),
    ("ML / AI ENGINEERING POD",      C_DEEP_BLUE, [
        ("Lead ML/AI Engineer",       "Offshore", [1]*24),
        ("Sr. ML/AI Engineer — RAG",  "Offshore", [0,0]+[1]*22),
        ("Sr. ML/AI Engineer — Agents","Offshore",[0,0]+[1]*22),
        ("ML/AI Engineer — Guardrails","Offshore",[0,0,0,0]+[1]*20),
        ("ML/AI Engineer — Fine-Tuning","Offshore",[0]*10+[1]*14),
    ]),
    ("DATA ENGINEERING POD",         C_TEAL, [
        ("Lead Data Engineer",        "Offshore", [1]*24),
        ("Sr. Data Engineer — Ingestion","Offshore",[1]*20+[0]*4),
        ("Data Engineer — Pipelines", "Offshore", [0,0]+[1]*18+[0]*4),
    ]),
    ("MLOPS & PLATFORM POD",         C_PURPLE, [
        ("Lead MLOps Engineer",       "Offshore", [1]*24),
        ("Sr. MLOps Engineer — Infra","Offshore", [1]*24),
        ("MLOps Engineer — CI/CD",    "Offshore", [0,0]+[1]*22),
    ]),
    ("QA & EVALUATION POD",          C_ORANGE, [
        ("AI QA Lead / Evaluation Eng","Offshore",[1]*24),
        ("AI QA Engineer — Functional","Offshore",[0,0,0,0]+[1]*20),
        ("Performance Test Engineer", "Offshore", [0]*16+[1]*8),
    ]),
    ("SECURITY & COMPLIANCE",        "5D6D7E", [
        ("AI Security Engineer",      "Offshore", [1,1,1,1]+[0]*16+[1,1,1,1]),
        ("PII / Compliance Specialist","Offshore",[1,1,1,1]+[0]*16+[1,1,1,1]),
    ]),
]

data_start_rl=9; rl_row=data_start_rl; shade_rl=False

for (pod_name, pod_color, resources) in POD_DATA:
    rl.row_dimensions[rl_row].height=20
    rl.merge_cells(f"A{rl_row}:{get_column_letter(26)}{rl_row}")
    ph=rl.cell(rl_row,1,f"  ▌  {pod_name}")
    ph.font=Font(name="Calibri",bold=True,color=C_WHITE,size=10)
    ph.fill=fill(pod_color); ph.alignment=aln("left","center",indent=1); ph.border=bdr()
    for c in range(2,27): rl.cell(rl_row,c).fill=fill(pod_color); rl.cell(rl_row,c).border=bdr()
    rl_row+=1

    for (role,location,weekly) in resources:
        rl.row_dimensions[rl_row].height=16
        bg=C_PALE_BLUE if shade_rl else "FFFFFF"; shade_rl=not shade_rl
        rc=rl.cell(rl_row,1,f"  {role}")
        rc.font=fnt(size=9,color=C_GREY_TXT); rc.fill=fill(bg)
        rc.alignment=aln("left","center",indent=2); rc.border=bdr()
        lc=rl.cell(rl_row,2,location)
        lc.font=fnt(size=9,color=C_GREY_TXT); lc.fill=fill(bg); lc.alignment=aln("center"); lc.border=bdr()
        for wi,cnt in enumerate(weekly):
            c=rl.cell(rl_row,wi+3,cnt if cnt>0 else None)
            c.fill=fill(C_PALE_GRN) if cnt>0 else fill(bg)
            c.font=Font(name="Calibri",bold=(cnt>0),size=9,color=C_NAVY if cnt>0 else "CCCCCC")
            c.alignment=aln("center"); c.border=bdr()
        rl_row+=1

# Totals
rl.row_dimensions[rl_row].height=22
tot=rl.cell(rl_row,1,"  WEEKLY TOTAL HEADCOUNT")
tot.font=Font(name="Calibri",bold=True,color=C_WHITE,size=10)
tot.fill=fill(C_NAVY); tot.alignment=aln("left","center",indent=1); tot.border=bdr()
rl.cell(rl_row,2).fill=fill(C_NAVY); rl.cell(rl_row,2).border=bdr()
for w in range(1,25):
    col=w+2
    c=rl.cell(rl_row,col,f"=SUM({get_column_letter(col)}{data_start_rl}:{get_column_letter(col)}{rl_row-1})")
    c.font=Font(name="Calibri",bold=True,color=C_WHITE,size=10)
    c.fill=fill(C_NAVY); c.alignment=aln("center"); c.border=bdr()

# ═══════════════════════════════════════════════════════════════════════
#  SHEET 3 — ESTIMATION SUMMARY
# ═══════════════════════════════════════════════════════════════════════
sm = wb.create_sheet("Estimation Summary")
sm.sheet_view.showGridLines = False
sm.column_dimensions["A"].width=38; sm.column_dimensions["B"].width=16
sm.column_dimensions["C"].width=16; sm.column_dimensions["D"].width=16
sm.column_dimensions["E"].width=16; sm.column_dimensions["F"].width=16

for r,h in [(1,8),(2,36),(3,16),(4,8)]: sm.row_dimensions[r].height=h
sm.merge_cells("A2:F2")
t=sm.cell(2,1,"EXPERIAN · AI ENGINEERING PHASE 1 — ESTIMATION SUMMARY")
t.font=Font(name="Calibri",bold=True,color=C_WHITE,size=15); t.fill=fill(C_NAVY); t.alignment=aln("center","center")
sm.merge_cells("A3:F3")
s=sm.cell(3,1,"Phase 1  |  AI-Only Team  |  June 2026  |  Offshore  |  9 Work Streams  |  Confidence: High")
s.font=Font(name="Calibri",italic=True,color="A9CCE3",size=10); s.fill=fill(C_NAVY); s.alignment=aln("center","center")
sm.merge_cells("A4:F4"); sm.cell(4,1).fill=fill(C_TEAL)
for c in range(2,7): sm.cell(2,c).fill=fill(C_NAVY); sm.cell(3,c).fill=fill(C_NAVY); sm.cell(4,c).fill=fill(C_TEAL)

# Headers
sm.row_dimensions[5].height=28
for ci,h in enumerate(["Work Stream","ML/AI Eng (PD)","Data Eng (PD)","MLOps (PD)","Total (PD)","Total (PM)"],1):
    c=sm.cell(5,ci,h)
    c.font=Font(name="Calibri",bold=True,color=C_WHITE,size=9)
    c.fill=fill(C_DEEP_BLUE); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=bdr()

SUM_DATA=[
    ("1.0  Intelligent Search",          57, 23, 21),
    ("2.0  Conversational Chatbot",       45, 17, 15),
    ("3.0  DocuSafe AI Insights",         67, 32, 20),
    ("4.0  Report Recommendations",       32, 10, 10),
    ("5.0  Knowledge Management & Help",  25,  11, 8),
    ("6.0  AEO / AI Readability",         17,   8, 8),
    ("7.0  AI Gateway & Shared Infra",    67,  37, 24),
    ("8.0  Platform Infrastructure",      14,  13, 39),
    ("9.0  NFR & Performance Engineering",26,  10, 12),
]

sm_row=6; shade_sm=False
for label,ml,de,mo in SUM_DATA:
    sm.row_dimensions[sm_row].height=18
    bg=C_PALE_BLUE if shade_sm else "FFFFFF"; shade_sm=not shade_sm
    c=sm.cell(sm_row,1,f"  {label}"); c.font=fnt(size=9,color=C_GREY_TXT); c.fill=fill(bg); c.alignment=aln("left","center",indent=1); c.border=bdr()
    for ci2,v in enumerate([ml,de,mo],2):
        c=sm.cell(sm_row,ci2,v); c.font=fnt(size=9,color="154360"); c.fill=fill(bg); c.alignment=aln("center"); c.border=bdr()
    tot=sm.cell(sm_row,5,f"=B{sm_row}+C{sm_row}+D{sm_row}")
    tot.font=Font(name="Calibri",bold=True,size=9,color=C_NAVY); tot.fill=fill(bg); tot.alignment=aln("center"); tot.border=bdr()
    pm=sm.cell(sm_row,6,f"=ROUND(E{sm_row}/21.5,1)")
    pm.font=Font(name="Calibri",bold=True,size=9,color=C_NAVY); pm.fill=fill(bg); pm.alignment=aln("center"); pm.border=bdr()
    sm_row+=1

# Overhead rows
OVERHEADS=[("Programme Governance (15%)","=ROUND(SUM(E6:E14)*0.15,0)"),
           ("Risk Contingency (15%)",    "=ROUND(SUM(E6:E14)*0.15,0)"),
           ("Warranty / Hypercare",      1500)]
for label,formula in OVERHEADS:
    sm.row_dimensions[sm_row].height=18
    c=sm.cell(sm_row,1,f"  {label}"); c.font=fnt(italic=True,size=9,color=C_GREY_TXT); c.fill=fill(C_SILVER); c.alignment=aln("left","center",indent=1); c.border=bdr()
    for ci2 in range(2,5): sm.cell(sm_row,ci2,"—").font=fnt(italic=True,size=9,color="AAAAAA"); sm.cell(sm_row,ci2).fill=fill(C_SILVER); sm.cell(sm_row,ci2).alignment=aln("center"); sm.cell(sm_row,ci2).border=bdr()
    tv=sm.cell(sm_row,5,formula); tv.font=Font(name="Calibri",bold=True,size=9,color=C_ORANGE); tv.fill=fill(C_SILVER); tv.alignment=aln("center"); tv.border=bdr()
    pv=sm.cell(sm_row,6,f"=ROUND(E{sm_row}/21.5,1)"); pv.font=Font(name="Calibri",bold=True,size=9,color=C_ORANGE); pv.fill=fill(C_SILVER); pv.alignment=aln("center"); pv.border=bdr()
    sm_row+=1

# Grand total
sm.row_dimensions[sm_row].height=24
sm.merge_cells(f"A{sm_row}:D{sm_row}")
gt=sm.cell(sm_row,1,"  GRAND TOTAL"); gt.font=Font(name="Calibri",bold=True,color=C_WHITE,size=11); gt.fill=fill(C_NAVY); gt.alignment=aln("left","center",indent=1); gt.border=bdr()
for c in range(2,5): sm.cell(sm_row,c).fill=fill(C_NAVY); sm.cell(sm_row,c).border=bdr()
gtv=sm.cell(sm_row,5,f"=SUM(E6:E{sm_row-1})"); gtv.font=Font(name="Calibri",bold=True,color=C_WHITE,size=12); gtv.fill=fill(C_NAVY); gtv.alignment=aln("center"); gtv.border=bdr()
gtp=sm.cell(sm_row,6,f"=ROUND(E{sm_row}/21.5,1)"); gtp.font=Font(name="Calibri",bold=True,color=C_WHITE,size=12); gtp.fill=fill(C_NAVY); gtp.alignment=aln("center"); gtp.border=bdr()
sm_row+=2

# Assumptions block
sm.row_dimensions[sm_row].height=18
sm.merge_cells(f"A{sm_row}:F{sm_row}")
ah=sm.cell(sm_row,1,"  KEY ASSUMPTIONS"); ah.font=Font(name="Calibri",bold=True,color=C_WHITE,size=10); ah.fill=fill(C_TEAL); ah.alignment=aln("left","center",indent=1); ah.border=bdr()
sm_row+=1
ASSUMPTIONS=[
    "AI-only delivery team: ML/AI Engineers, Data Engineers, MLOps Engineers, AI QA, Security. No FE/BE developers.",
    "8 hrs/day | 21.5 days/month | 172 hrs/month — all Offshore (India).",
    "Primary LLM: AWS Bedrock (Anthropic Claude). Fallback: SageMaker OSS (Llama 3/Mistral) → Azure OpenAI.",
    "OpenSearch (hybrid dense+sparse) as Vector Store; Aurora pgvector for metadata/embeddings.",
    "PII tokenisation vault is MY-resident; raw PII never crosses MY→SG border (PDPA compliant).",
    "LangGraph used for agent orchestrator; MCP protocol for tool-server integration.",
    "Gen AI productivity uplift: ML/AI 15–35% of base estimates (code gen + prompt iteration acceleration).",
    "RAGAS evaluation suite baseline per use-case: faithfulness > 0.85, context precision > 0.80.",
    "Risk contingency 15% | Programme governance 15% | Warranty: 1,500 person-hours post go-live.",
    "Continuous learning loop: telemetry → golden set curation → SageMaker LoRA/QLoRA fine-tune.",
]
for asm in ASSUMPTIONS:
    sm.row_dimensions[sm_row].height=16
    sm.merge_cells(f"A{sm_row}:F{sm_row}")
    ac=sm.cell(sm_row,1,f"  •  {asm}"); ac.font=fnt(size=9,color=C_GREY_TXT); ac.fill=fill(C_SILVER); ac.alignment=aln("left","center",indent=1,wrap=True); ac.border=bdr()
    sm_row+=1

wb.move_sheet("Estimation Summary",offset=-2)

# SAVE (JD sheet added next)
out="Experian_AI_Phase1_WBS.xlsx"
wb.save(out); print("WBS + RL + Summary saved. Rows WBS:", row, "RL:", rl_row)
