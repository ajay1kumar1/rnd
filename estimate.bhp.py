from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

wb = Workbook()
print("test")
# ── helpers ──────────────────────────────────────────────────────────────────
def rgb(hex_str):
    return hex_str.lstrip("#")

def hdr_fill(hex_str):   return PatternFill("solid", fgColor=rgb(hex_str))
def hdr_font(hex_str="FFFFFF", sz=11, bold=True, name="Arial"):
    return Font(name=name, bold=bold, color=rgb(hex_str), size=sz)
def body_font(sz=10, bold=False, color="1A1A2E", name="Arial"):
    return Font(name=name, bold=bold, color=rgb(color), size=sz)
def center():  return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():    return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def right():   return Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bottom():
    s   = Side(style="medium", color="028090")
    thn = Side(style="thin",   color="CCCCCC")
    return Border(left=thn, right=thn, top=thn, bottom=s)

# colour palette
C_NAVY  = "0D1B3E"   # header bg
C_TEAL  = "028090"   # section header
C_MINT  = "02C39A"   # accent / sub-header
C_LGRAY = "F2F6FA"   # alternate row
C_WHITE = "FFFFFF"
C_YELL  = "FFF9C4"   # total row highlight
C_DRED  = "C62828"   # risk high
C_DGRN  = "1B5E20"   # risk low

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 – API Estimation Detail
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "API Estimation Detail"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A5"

# ── col widths ──
col_widths = {
    "A": 5,   # #
    "B": 28,  # Module
    "C": 36,  # API / Task
    "D": 14,  # Method
    "E": 12,  # Best (hrs)
    "F": 12,  # Likely (hrs)
    "G": 12,  # Worst (hrs)
    "H": 14,  # PERT Est (hrs)
    "I": 12,  # Days (8h)
    "J": 14,  # Complexity
    "K": 28,  # Notes / Dependencies
}
for col, w in col_widths.items():
    ws1.column_dimensions[col].width = w

# row 1 – title banner
ws1.row_dimensions[1].height = 36
ws1.merge_cells("A1:K1")
ws1["A1"] = "API Development Time Estimation  ·  Scope of Work"
ws1["A1"].font   = Font(name="Arial", bold=True, size=16, color=rgb(C_WHITE))
ws1["A1"].fill   = hdr_fill(C_NAVY)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

# row 2 – sub-banner
ws1.row_dimensions[2].height = 22
ws1.merge_cells("A2:K2")
ws1["A2"] = "Stack: Python · FastAPI · Microsoft Graph API · Unifonic SMS · JWT · OAuth 2.0 / OIDC"
ws1["A2"].font = Font(name="Arial", size=10, italic=True, color=rgb(C_WHITE))
ws1["A2"].fill = hdr_fill(C_TEAL)
ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")

# row 3 – PERT formula note
ws1.row_dimensions[3].height = 18
ws1.merge_cells("A3:K3")
ws1["A3"] = "PERT Estimate = (Best + 4 × Likely + Worst) / 6     |     Days = PERT Hours ÷ 8"
ws1["A3"].font = Font(name="Arial", size=9, italic=True, color=rgb("555555"))
ws1["A3"].alignment = Alignment(horizontal="center", vertical="center")

# row 4 – column headers
ws1.row_dimensions[4].height = 30
headers = ["#", "Module", "API / Task", "HTTP Method",
           "Best\n(hrs)", "Likely\n(hrs)", "Worst\n(hrs)",
           "PERT Est\n(hrs)", "Days\n(8h)", "Complexity", "Notes / Dependencies"]
for ci, h in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=ci, value=h)
    cell.font      = hdr_font(sz=10)
    cell.fill      = hdr_fill(C_NAVY)
    cell.alignment = center()
    cell.border    = thin_border()

# ── data rows ──
# columns: (Module, API/Task, Method, Best, Likely, Worst, Complexity, Notes)
rows_data = [
    # ── 1. User Registration ──
    ("__SECTION__", "1. USER REGISTRATION APIs", "", "", "", "", "", "", "", "", ""),
    (1,  "User Registration", "Register via Email",               "POST",   2,  4,  6,  "", "", "Medium",  "Validate email format, check duplicate, hash password, generate verification token"),
    (2,  "User Registration", "Register via Mobile Number",        "POST",   2,  4,  6,  "", "", "Medium",  "E.164 format validation, duplicate check, trigger OTP via Unifonic"),
    (3,  "User Registration", "Verify OTP – Registration",         "POST",   1,  2,  4,  "", "", "Low",     "Validate OTP code, expiry check, mark account active"),
    (4,  "User Registration", "Resend OTP",                        "POST",   1,  2,  3,  "", "", "Low",     "Rate-limit resend attempts, regenerate token, re-trigger Unifonic"),
    # ── 2. Unifonic Integration ──
    ("__SECTION__", "2. UNIFONIC SMS INTEGRATION", "", "", "", "", "", "", "", "", ""),
    (5,  "Unifonic SMS",      "Send OTP – Registration",           "POST",   2,  3,  5,  "", "", "Medium",  "Unifonic REST API wrapper; template management, retry on failure"),
    (6,  "Unifonic SMS",      "Send OTP – Login",                  "POST",   1,  2,  3,  "", "", "Low",     "Reuse Unifonic service; login-specific template"),
    (7,  "Unifonic SMS",      "Send OTP – Password Reset",         "POST",   1,  2,  3,  "", "", "Low",     "Password reset template; link expiry in message body"),
    (8,  "Unifonic SMS",      "OTP Validation & Expiry Mgmt",      "–",      2,  4,  6,  "", "", "Medium",  "Time-based OTP validity (5 min), single-use enforcement, Redis/DB store"),
    # ── 3. Microsoft Graph API ──
    ("__SECTION__", "3. MICROSOFT GRAPH API INTEGRATION", "", "", "", "", "", "", "", "", ""),
    (9,  "MS Graph API",      "Create User in Entra External ID",  "POST",   3,  5,  8,  "", "", "High",    "Graph API auth (client credentials), map registration payload to Entra schema"),
    (10, "MS Graph API",      "Update User Profile in Entra",      "PATCH",  2,  4,  6,  "", "", "Medium",  "Delta update only changed attributes; handle Entra rate limits"),
    (11, "MS Graph API",      "Retrieve User Information",         "GET",    1,  3,  5,  "", "", "Medium",  "Fetch by objectId or email; map Entra attributes to local model"),
    (12, "MS Graph API",      "Manage Custom User Attributes",     "PATCH",  3,  5,  8,  "", "", "High",    "Define extension attributes in Entra; read/write via Graph extensions endpoint"),
    (13, "MS Graph API",      "Graph API Auth Module (service)",   "–",      3,  5,  7,  "", "", "High",    "OAuth 2.0 client_credentials flow; token cache; refresh logic; unit tests"),
    # ── 4. Profile Management ──
    ("__SECTION__", "4. PROFILE MANAGEMENT APIs", "", "", "", "", "", "", "", "", ""),
    (14, "Profile Mgmt",      "View Profile",                      "GET",    1,  2,  3,  "", "", "Low",     "Return profile from local DB + Entra merge; mask sensitive fields"),
    (15, "Profile Mgmt",      "Update Profile",                    "PATCH",  2,  4,  6,  "", "", "Medium",  "Validate fields (DOB, mobile format, country code), sync to Entra via Graph"),
    # ── 5. User Migration ──
    ("__SECTION__", "5. USER MIGRATION", "", "", "", "", "", "", "", "", ""),
    (16, "User Migration",    "CSV Ingestion & Validation Script", "–",      3,  5,  8,  "", "", "High",    "pandas-based reader; schema validation; duplicate detection; error report CSV"),
    (17, "User Migration",    "Data Transformation Layer",        "–",      2,  4,  6,  "", "", "Medium",  "Map CSV columns to Entra schema; handle nulls, encoding, phone normalisation"),
    (18, "User Migration",    "Bulk User Creation via Graph API",  "POST",   4,  6, 10,  "", "", "High",    "Batch Graph calls (≤20/req); throttle handling; idempotency; progress log"),
    (19, "User Migration",    "Password Reset / Temp Password",    "POST",   2,  3,  5,  "", "", "Medium",  "Assign temp password; force-reset flag in Entra; trigger welcome SMS/email"),
    (20, "User Migration",    "Migration Error Handling & Retry",  "–",      2,  4,  6,  "", "", "Medium",  "Retry queue for failed records; partial success report; audit trail"),
    # ── 6. Authentication & Security ──
    ("__SECTION__", "6. AUTHENTICATION & SECURITY", "", "", "", "", "", "", "", "", ""),
    (21, "Auth & Security",   "Login – Email + Password",          "POST",   2,  4,  6,  "", "", "Medium",  "Verify credentials, issue JWT (access + refresh tokens), log event"),
    (22, "Auth & Security",   "Login – Mobile OTP",                "POST",   2,  3,  5,  "", "", "Medium",  "Trigger Unifonic OTP; validate; issue JWT on success"),
    (23, "Auth & Security",   "Password Reset – Request OTP",      "POST",   1,  2,  4,  "", "", "Low",     "Lookup account by email/mobile; trigger Unifonic reset OTP"),
    (24, "Auth & Security",   "Password Reset – Set New Password", "POST",   2,  3,  5,  "", "", "Medium",  "Validate OTP, enforce password policy, update in Entra + local DB"),
    (25, "Auth & Security",   "Token Refresh",                     "POST",   1,  2,  3,  "", "", "Low",     "Validate refresh token, rotate & issue new pair, revoke old"),
    (26, "Auth & Security",   "Logout / Token Revocation",         "POST",   1,  2,  3,  "", "", "Low",     "Blacklist token in Redis/DB; clear session state"),
    (27, "Auth & Security",   "JWT Middleware (FastAPI Dependency)","–",      2,  4,  6,  "", "", "Medium",  "Decode & verify JWT on every protected route; role/scope extraction"),
    (28, "Auth & Security",   "Error Handling & Logging Infra",    "–",      2,  4,  6,  "", "", "Medium",  "Structured logging (JSON), error codes, Sentry/CloudWatch hook, stack traces"),
    # ── 7. Cross-cutting / Infra ──
    ("__SECTION__", "7. CROSS-CUTTING & INFRASTRUCTURE", "", "", "", "", "", "", "", "", ""),
    (29, "Infrastructure",    "FastAPI Project Scaffold & Config", "–",      2,  4,  6,  "", "", "Medium",  "Project structure, .env config, DB connection pool, CORS, docs (OpenAPI)"),
    (30, "Infrastructure",    "Database Schema & Migrations",      "–",      2,  4,  6,  "", "", "Medium",  "SQLAlchemy models / Alembic migrations for users, OTP, audit tables"),
    (31, "Infrastructure",    "Unit & Integration Test Suite",     "–",      4,  8, 12,  "", "", "High",    "pytest + httpx; mock Graph API & Unifonic; ≥80% coverage target"),
    (32, "Infrastructure",    "API Documentation (OpenAPI/Swagger)","–",     1,  2,  3,  "", "", "Low",     "Auto-generated via FastAPI; manual descriptions for all endpoints"),
    (33, "Infrastructure",    "Deployment Config (Docker / CI)",   "–",      2,  4,  6,  "", "", "Medium",  "Dockerfile, docker-compose, GitHub Actions CI pipeline, env secrets"),
]

COMPLEXITY_COLOR = {
    "Low":    "E8F5E9",
    "Medium": "FFF8E1",
    "High":   "FFEBEE",
}

ROW_START = 5
data_rows = []   # (excel_row, is_section, data_tuple)
r = ROW_START

for item in rows_data:
    if item[0] == "__SECTION__":
        data_rows.append((r, True, item))
    else:
        data_rows.append((r, False, item))
    r += 1

# write rows
pert_cells  = []   # (row, col_H) to write PERT formula
days_cells  = []   # (row, col_I)
data_row_refs = [] # (row) for each non-section data row

for (excel_row, is_section, item) in data_rows:
    ws1.row_dimensions[excel_row].height = 22

    if is_section:
        ws1.merge_cells(f"A{excel_row}:K{excel_row}")
        cell = ws1[f"A{excel_row}"]
        cell.value     = item[1]
        cell.font      = Font(name="Arial", bold=True, size=11, color=rgb(C_WHITE))
        cell.fill      = hdr_fill(C_TEAL)
        cell.alignment = left()
        cell.border    = thick_bottom()
    else:
        seq, module, task, method, best, likely, worst, _, _, complexity, notes = item
        vals = [seq, module, task, method, best, likely, worst, "", "", complexity, notes]
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=excel_row, column=ci, value=v)
            row_bg = C_WHITE if (excel_row % 2 == 0) else C_LGRAY
            cell.fill   = hdr_fill(row_bg)
            cell.border = thin_border()

            if ci in (1,):        # #
                cell.alignment = center()
                cell.font = body_font(bold=True, color=C_TEAL)
            elif ci in (4,):      # method
                cell.alignment = center()
                cell.font = body_font(bold=True, color="5C6BC0")
            elif ci in (5, 6, 7): # est hours
                cell.alignment = center()
                cell.font = body_font(bold=True, color="1A237E")
                cell.number_format = "0.0"
            elif ci == 8:         # PERT
                E = get_column_letter(5)
                F = get_column_letter(6)
                G = get_column_letter(7)
                cell.value = f"=({E}{excel_row}+4*{F}{excel_row}+{G}{excel_row})/6"
                cell.font  = body_font(bold=True, color=C_TEAL)
                cell.alignment = center()
                cell.number_format = "0.0"
                cell.fill = hdr_fill("E0F7FA")
            elif ci == 9:         # Days
                cell.value = f"=H{excel_row}/8"
                cell.font  = body_font(bold=True, color="880E4F")
                cell.alignment = center()
                cell.number_format = "0.0"
                cell.fill = hdr_fill("FCE4EC")
            elif ci == 10:        # Complexity
                cell.alignment = center()
                comp_bg = COMPLEXITY_COLOR.get(str(v), C_WHITE)
                cell.fill = hdr_fill(comp_bg)
                comp_color = {"Low": C_DGRN, "Medium": "E65100", "High": C_DRED}.get(str(v), "000000")
                cell.font = Font(name="Arial", bold=True, size=10, color=rgb(comp_color))
            elif ci == 11:        # Notes
                cell.font = Font(name="Arial", size=9, italic=True, color=rgb("455A64"))
                cell.alignment = left()
            else:
                cell.alignment = left()
                cell.font = body_font()

        data_row_refs.append(excel_row)

# ── TOTALS ROW ──
total_row = r
ws1.row_dimensions[total_row].height = 28
ws1.merge_cells(f"A{total_row}:D{total_row}")
tc = ws1[f"A{total_row}"]
tc.value = "TOTAL  (33 APIs / Tasks)"
tc.font  = Font(name="Arial", bold=True, size=11, color=rgb(C_WHITE))
tc.fill  = hdr_fill(C_NAVY)
tc.alignment = center()
tc.border = thick_bottom()

for col_idx, col_letter in enumerate(["E", "F", "G", "H", "I"], 5):
    cell = ws1.cell(row=total_row, column=col_idx)
    # sum only data rows (skip section rows)
    refs = "+".join([f"{col_letter}{rw}" for rw in data_row_refs])
    cell.value  = f"={refs}"
    cell.font   = Font(name="Arial", bold=True, size=12, color=rgb(C_NAVY))
    cell.fill   = hdr_fill(C_YELL)
    cell.alignment = center()
    cell.number_format = "0.0"
    cell.border = thick_bottom()

for col_idx in [10, 11]:
    cell = ws1.cell(row=total_row, column=col_idx)
    cell.fill   = hdr_fill(C_YELL)
    cell.border = thick_bottom()

# empty col 10 total label
ws1.cell(row=total_row, column=10).value = ""
ws1.cell(row=total_row, column=11).value = "See Summary sheet for phase breakdown"
ws1.cell(row=total_row, column=11).font = Font(name="Arial", italic=True, size=9, color=rgb("555555"))

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 – Project Summary
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Project Summary")
ws2.sheet_view.showGridLines = False

col_widths2 = {"A": 6, "B": 34, "C": 14, "D": 14, "E": 14, "F": 14, "G": 28}
for col, w in col_widths2.items():
    ws2.column_dimensions[col].width = w

# title
ws2.row_dimensions[1].height = 36
ws2.merge_cells("A1:G1")
ws2["A1"] = "Project Summary & Effort Breakdown"
ws2["A1"].font      = Font(name="Arial", bold=True, size=16, color=rgb(C_WHITE))
ws2["A1"].fill      = hdr_fill(C_NAVY)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws2.row_dimensions[2].height = 20
ws2.merge_cells("A2:G2")
ws2["A2"] = "Estimates use PERT method (3-point): Best / Likely / Worst  ·  1 day = 8 hrs  ·  1 sprint = 10 days"
ws2["A2"].font = Font(name="Arial", size=10, italic=True, color=rgb(C_WHITE))
ws2["A2"].fill = hdr_fill(C_TEAL)
ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")

# ── Phase table ──
phase_hdr_row = 4
ws2.row_dimensions[phase_hdr_row].height = 28
p_hdrs = ["#", "Module / Phase", "APIs\nCount", "Est Hours\n(PERT)", "Days\n(÷8)", "Sprints\n(÷10)", "Key Deliverables"]
for ci, h in enumerate(p_hdrs, 1):
    c = ws2.cell(row=phase_hdr_row, column=ci, value=h)
    c.font = hdr_font(sz=10)
    c.fill = hdr_fill(C_NAVY)
    c.alignment = center()
    c.border = thin_border()

phase_data = [
    (1, "User Registration APIs",          4,   44,  "Medium", "Register (Email+Mobile), OTP verify, Resend OTP"),
    (2, "Unifonic SMS Integration",         4,   44,  "Medium", "OTP send wrappers, validation, expiry management"),
    (3, "Microsoft Graph API Integration",  5,   72,  "High",   "User CRUD in Entra, custom attributes, auth module"),
    (4, "Profile Management APIs",          2,   22,  "Low",    "View + Update profile, Entra sync"),
    (5, "User Migration",                   5,   76,  "High",   "CSV ingest, transform, bulk create, password reset"),
    (6, "Authentication & Security",        8,   82,  "High",   "Login flows, JWT, OTP reset, middleware, logging"),
    (7, "Cross-cutting & Infrastructure",   5,   66,  "High",   "Scaffold, DB schema, tests, docs, Docker/CI"),
]

pr = phase_hdr_row + 1
phase_row_refs = []
for item in phase_data:
    ws2.row_dimensions[pr].height = 22
    num, module, count, est_hrs, comp, deliverables = item
    bg = C_WHITE if pr % 2 == 0 else C_LGRAY
    for ci, v in enumerate([num, module, count, est_hrs, "", "", deliverables], 1):
        c = ws2.cell(row=pr, column=ci, value=v)
        c.fill   = hdr_fill(bg)
        c.border = thin_border()
        if ci == 1:
            c.font = body_font(bold=True, color=C_TEAL); c.alignment = center()
        elif ci == 3:
            c.font = body_font(bold=True); c.alignment = center()
        elif ci == 4:
            c.value = est_hrs
            c.font = Font(name="Arial", bold=True, size=10, color=rgb("1A237E"))
            c.alignment = center(); c.number_format = "0.0"
        elif ci == 5:   # days formula
            c.value = f"=D{pr}/8"
            c.font = Font(name="Arial", bold=True, size=10, color=rgb("880E4F"))
            c.alignment = center(); c.number_format = "0.0"
            c.fill = hdr_fill("FCE4EC")
        elif ci == 6:   # sprints formula
            c.value = f"=E{pr}/10"
            c.font = Font(name="Arial", bold=True, size=10, color=rgb(C_TEAL))
            c.alignment = center(); c.number_format = "0.0"
            c.fill = hdr_fill("E0F7FA")
        elif ci == 7:
            c.font = Font(name="Arial", size=9, italic=True, color=rgb("455A64"))
            c.alignment = left()
        else:
            c.font = body_font(); c.alignment = left()
    phase_row_refs.append(pr)
    pr += 1

# totals
ws2.row_dimensions[pr].height = 28
ws2.merge_cells(f"A{pr}:B{pr}")
tc = ws2[f"A{pr}"]
tc.value = "GRAND TOTAL  (33 APIs / Tasks)"
tc.font  = Font(name="Arial", bold=True, size=11, color=rgb(C_WHITE))
tc.fill  = hdr_fill(C_NAVY)
tc.alignment = center()
tc.border = thick_bottom()

for col_idx, col_letter in enumerate(["C", "D", "E", "F"], 3):
    c = ws2.cell(row=pr, column=col_idx)
    refs = "+".join([f"{col_letter}{rw}" for rw in phase_row_refs])
    c.value = f"={refs}"
    c.font  = Font(name="Arial", bold=True, size=12, color=rgb(C_NAVY))
    c.fill  = hdr_fill(C_YELL)
    c.alignment = center()
    c.number_format = "0.0"
    c.border = thick_bottom()

ws2.cell(row=pr, column=7).fill   = hdr_fill(C_YELL)
ws2.cell(row=pr, column=7).border = thick_bottom()

# ── Timeline / Assumptions block ──
ta_start = pr + 2
ws2.row_dimensions[ta_start].height = 26
ws2.merge_cells(f"A{ta_start}:G{ta_start}")
c = ws2[f"A{ta_start}"]
c.value = "DELIVERY ASSUMPTIONS & TIMELINE PARAMETERS"
c.font  = Font(name="Arial", bold=True, size=11, color=rgb(C_WHITE))
c.fill  = hdr_fill(C_TEAL)
c.alignment = Alignment(horizontal="left", vertical="center")

assumptions = [
    ("Working Hours per Day",         "8 hrs",        "Standard working day"),
    ("Sprint Duration",                "10 days",      "2-week sprints"),
    ("Team Velocity (per sprint)",     "1 developer",  "Single dev assigned to this scope"),
    ("Estimated Sprints (Phase 1)",    "5–6 sprints",  "~50–60 days for core scope"),
    ("Estimated Calendar Duration",    "10–12 weeks",  "Accounting for reviews, QA, integration lag"),
    ("Buffer / Risk Contingency",      "20%",          "Added on top of PERT worst-case for infra & Graph complexity"),
    ("Recommended Team",              "1 Senior Dev + 1 QA", "Senior Python/FastAPI dev · QA for test suite"),
    ("Environment Setup",              "~2–3 days",    "Entra external ID tenant setup, Unifonic account, Azure app reg"),
    ("Not included in estimate",       "DevOps / Infra provisioning beyond Docker", "AWS/Azure setup, Keycloak, full CI pipeline cost"),
]

for i, (param, value, note) in enumerate(assumptions):
    row_n = ta_start + 1 + i
    ws2.row_dimensions[row_n].height = 20
    bg = C_WHITE if i % 2 == 0 else C_LGRAY
    for ci, v in enumerate([" " + param, value, note], 1):
        c = ws2.cell(row=row_n, column=ci)
        c.fill   = hdr_fill(bg)
        c.border = thin_border()
        if ci == 1:
            c.font = Font(name="Arial", bold=True, size=10, color=rgb(C_NAVY))
            c.alignment = left()
            ws2.merge_cells(f"A{row_n}:B{row_n}")
        elif ci == 2:
            c.font = Font(name="Arial", bold=True, size=10, color=rgb(C_TEAL))
            c.alignment = center()
            ws2.merge_cells(f"C{row_n}:D{row_n}")
        elif ci == 3:
            c.font = Font(name="Arial", size=9, italic=True, color=rgb("455A64"))
            c.alignment = left()
            ws2.merge_cells(f"E{row_n}:G{row_n}")

# ── Risk / Complexity Legend ──
leg_start = ta_start + len(assumptions) + 2
ws2.row_dimensions[leg_start].height = 26
ws2.merge_cells(f"A{leg_start}:G{leg_start}")
c = ws2[f"A{leg_start}"]
c.value = "COMPLEXITY & RISK LEGEND"
c.font  = Font(name="Arial", bold=True, size=11, color=rgb(C_WHITE))
c.fill  = hdr_fill(C_NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")

legend = [
    ("Low",    "E8F5E9", C_DGRN,  "Simple CRUD, single service call, minimal validation"),
    ("Medium", "FFF8E1", "E65100","2+ service dependencies, async, validation logic"),
    ("High",   "FFEBEE", C_DRED,  "External API + auth flow, batch processing, migrations, test coverage required"),
]
for i, (label, bg, fc, desc) in enumerate(legend):
    row_n = leg_start + 1 + i
    ws2.row_dimensions[row_n].height = 20
    for ci in range(1, 8):
        c = ws2.cell(row=row_n, column=ci)
        c.fill = hdr_fill(bg); c.border = thin_border()
        if ci == 1:
            c.value = label
            c.font  = Font(name="Arial", bold=True, size=10, color=rgb(fc))
            c.alignment = center()
            ws2.merge_cells(f"A{row_n}:B{row_n}")
        elif ci == 3:
            c.value = desc
            c.font  = Font(name="Arial", size=10, color=rgb("1A1A2E"))
            c.alignment = left()
            ws2.merge_cells(f"C{row_n}:G{row_n}")

# ═══════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════
out = "SOW_API_Time_Estimate.xlsx"
wb.save(out)
print("Saved:", out)